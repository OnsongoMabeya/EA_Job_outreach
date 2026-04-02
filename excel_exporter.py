"""
Excel Exporter Module
Exports job data to Excel format with proper formatting and text wrapping
"""

import json
import logging
from pathlib import Path
from datetime import datetime
from config import (
    VALIDATED_JOBS_FILE,
    SHEET_HEADERS
)

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("openpyxl not installed. Excel export will not be available.")


class ExcelExporter:
    """Exports job data to Excel file with formatting"""
    
    def __init__(self, output_file=None):
        """
        Initialize Excel exporter
        
        Args:
            output_file: Path to output Excel file (default: ea_jobs_YYYYMMDD_HHMMSS.xlsx)
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        
        if output_file is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"ea_jobs_{timestamp}.xlsx"
        
        self.output_file = Path(output_file)
        logger.info(f"Initialized Excel exporter: {self.output_file}")
    
    def load_from_file(self, filepath=VALIDATED_JOBS_FILE):
        """Load jobs from JSON file"""
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                jobs = json.load(f)
            logger.info(f"Loaded {len(jobs)} jobs from {filepath}")
            return jobs
        except FileNotFoundError:
            logger.error(f"File not found: {filepath}")
            return []
        except json.JSONDecodeError as e:
            logger.error(f"Invalid JSON in {filepath}: {e}")
            return []
    
    def export_jobs(self, jobs):
        """
        Export jobs to Excel file with formatting
        
        Args:
            jobs: List of job dictionaries
            
        Returns:
            dict with export statistics
        """
        if not jobs:
            logger.warning("No jobs to export")
            return {
                'exported': 0,
                'total': 0,
                'file': str(self.output_file)
            }
        
        logger.info(f"Exporting {len(jobs)} jobs to Excel...")
        
        try:
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "EA Jobs"
            
            # Write headers with formatting
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            for col_num, header in enumerate(SHEET_HEADERS, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Write data rows
            for row_num, job in enumerate(jobs, 2):
                # Truncate description if too long
                description = job.get('job_description', 'N/A')
                if len(description) > 500:
                    description = description[:497] + "..."
                
                row_data = [
                    job.get('job_url', 'N/A'),
                    job.get('job_title', 'N/A'),
                    job.get('company_name', 'N/A'),
                    job.get('location', 'Remote'),
                    description,
                    job.get('outreach_message', 'N/A')
                ]
                
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    # Enable text wrapping for all cells
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Set column widths
            column_widths = {
                'A': 50,  # Job Posting URL
                'B': 30,  # Job Title
                'C': 25,  # Company Name
                'D': 20,  # Location
                'E': 60,  # Job Description
                'F': 50   # LLM-Generated Outreach Message
            }
            
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width
            
            # Freeze header row
            ws.freeze_panes = 'A2'
            
            # Save workbook
            wb.save(self.output_file)
            
            logger.info(f"✓ Successfully exported {len(jobs)} jobs to {self.output_file}")
            
            return {
                'exported': len(jobs),
                'total': len(jobs),
                'file': str(self.output_file.absolute())
            }
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            raise
    
    def get_file_path(self):
        """Get the absolute path to the exported Excel file"""
        return str(self.output_file.absolute())


def main():
    """Test function for Excel exporter module"""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    
    if not EXCEL_AVAILABLE:
        print("\n❌ Excel export requires openpyxl package")
        print("Install with: pip install openpyxl\n")
        return
    
    exporter = ExcelExporter()
    jobs = exporter.load_from_file()
    
    if jobs:
        result = exporter.export_jobs(jobs)
        
        print(f"\n{'='*60}")
        print(f"EXCEL EXPORT COMPLETE")
        print(f"{'='*60}")
        print(f"✓ Exported: {result['exported']} jobs")
        print(f"Total processed: {result['total']} jobs")
        print(f"\nFile location: {result['file']}")
        print(f"\nThe Excel file has:")
        print(f"  • Text wrapping enabled for all cells")
        print(f"  • Formatted headers with colors")
        print(f"  • Optimized column widths")
        print(f"  • Frozen header row")
        print(f"{'='*60}\n")
    else:
        print("No jobs found. Run the full pipeline first.")


if __name__ == "__main__":
    main()
